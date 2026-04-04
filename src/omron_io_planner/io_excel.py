# -*- coding: utf-8 -*-
"""Excel 导入导出：支持欧姆龙风格表头（名称、数据类型、地址/值、注释等）。"""
from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Optional, Sequence

from openpyxl import Workbook, load_workbook

from .addressing import parse_cio_bit
from .models import IoPoint, IoProject, project_single_channel
from .omron_symbol_types import normalize_data_type

# 表头别名（小写匹配）
_COL_NAME = ("名称", "name", "符号", "符号名", "变量名", "symbol")
_COL_DTYPE = ("数据类型", "datatype", "类型", "type")
_COL_ADDR = ("地址/值", "地址", "address", "addr", "值", "变量地址", "io")
_COL_COMMENT = ("注释", "comment", "说明", "描述")
_COL_RACK = ("机架位置", "机架", "rack", "分组", "group", "槽")
_COL_USAGE = ("使用", "usage", "用途")
# 旧版导入
_COL_DIR = ("方向", "direction", "inout", "i/o")
_COL_SYM = ("符号", "symbol")


def _norm_header(cell) -> str:
    if cell is None:
        return ""
    return str(cell).strip().lower()


def _match_col(header: str, aliases: Sequence[str]) -> bool:
    h = header.strip().lower()
    return h in {a.lower() for a in aliases}


def _map_headers(row: Sequence) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for i, cell in enumerate(row):
        h = _norm_header(cell)
        if not h:
            continue
        if _match_col(h, _COL_NAME) and "name" not in mapping:
            mapping["name"] = i
        elif _match_col(h, _COL_DTYPE) and "data_type" not in mapping:
            mapping["data_type"] = i
        elif _match_col(h, _COL_ADDR) and "address" not in mapping:
            mapping["address"] = i
        elif _match_col(h, _COL_COMMENT) and "comment" not in mapping:
            mapping["comment"] = i
        elif _match_col(h, _COL_RACK) and "rack" not in mapping:
            mapping["rack"] = i
        elif _match_col(h, _COL_USAGE) and "usage" not in mapping:
            mapping["usage"] = i
        elif _match_col(h, _COL_DIR) and "direction" not in mapping:
            mapping["direction"] = i
        elif _match_col(h, _COL_SYM) and "symbol" not in mapping:
            mapping["symbol"] = i
    return mapping


def _has_omron_header(mapping: Dict[str, int]) -> bool:
    return "address" in mapping or "name" in mapping


def _looks_like_legacy(rows: List) -> bool:
    for row in rows[:80]:
        if not row:
            continue
        u = _cell_str(row[0]).upper()
        if u in ("IN", "OUT"):
            return True
    return False


def _addr_from_cell(raw) -> str:
    n = _norm_addr(raw)
    if n:
        return n
    return _cell_str(raw)


def _norm_addr(v) -> Optional[str]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        if isinstance(v, float) and abs(v - round(v, 2)) < 1e-9:
            s = f"{v:.2f}"
        else:
            s = str(v)
    else:
        s = str(v).strip()
    if not s:
        return None
    t = parse_cio_bit(s)
    if t:
        w, b = t
        return f"{w}.{b:02d}"
    return s


def _cell_str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _get_cell(row: Sequence, idx: Optional[int]) -> str:
    if idx is None or idx >= len(row):
        return ""
    return _cell_str(row[idx])


def _parse_flat_rows(name: str, rows: List, mapping: Dict[str, int]) -> IoProject:
    points: List[IoPoint] = []
    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        addr = ""
        if mapping.get("address") is not None and mapping["address"] < len(row):
            addr = _addr_from_cell(row[mapping["address"]])
        nm = _get_cell(row, mapping.get("name"))
        if not nm and mapping.get("symbol") is not None:
            nm = _get_cell(row, mapping["symbol"])
        dtype_raw = _get_cell(row, mapping.get("data_type"))
        dtype = normalize_data_type(dtype_raw or "")
        comment = _get_cell(row, mapping.get("comment"))
        rack = _get_cell(row, mapping.get("rack"))
        usage = _get_cell(row, mapping.get("usage"))
        if mapping.get("direction") is not None:
            d = _get_cell(row, mapping.get("direction"))
            if d and not usage:
                usage = d.upper() if d.upper() in ("IN", "OUT") else d
        if not addr.strip() and not nm.strip():
            continue
        points.append(
            IoPoint(
                name=nm,
                data_type=dtype,
                address=addr.strip(),
                comment=comment,
                rack=rack,
                usage=usage,
            )
        )
    return project_single_channel(name, "导入", points)


def _parse_fixed_columns(name: str, rows: List) -> IoProject:
    """无表头：名称, 数据类型, 地址/值, 注释, 机架位置, 使用。"""
    points: List[IoPoint] = []
    for row in rows:
        if not row:
            continue
        nm = _get_cell(row, 0)
        dtype = normalize_data_type(_get_cell(row, 1)) if len(row) > 1 else "BOOL"
        addr = _addr_from_cell(row[2]) if len(row) > 2 else ""
        if not addr:
            addr = _get_cell(row, 2) if len(row) > 2 else ""
        comment = _get_cell(row, 3) if len(row) > 3 else ""
        rack = _get_cell(row, 4) if len(row) > 4 else ""
        usage = _get_cell(row, 5) if len(row) > 5 else ""
        if not str(addr).strip() and not nm.strip():
            continue
        points.append(
            IoPoint(
                name=nm,
                data_type=dtype,
                address=str(addr).strip(),
                comment=comment,
                rack=rack,
                usage=usage,
            )
        )
    return project_single_channel(name, "导入", points)


def import_flat_table(
    path: str | Path,
    sheet_name: Optional[str] = None,
    sheet_index: int = 0,
) -> IoProject:
    path = Path(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError(f"工作表不存在: {sheet_name}，现有: {wb.sheetnames}")
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[sheet_index]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return project_single_channel(path.stem, "通道1", [])

    mapping = _map_headers(rows[0])
    if _has_omron_header(mapping):
        return _parse_flat_rows(path.stem, rows, mapping)
    return _parse_fixed_columns(path.stem, rows)


def import_legacy_blocks(path: str | Path, sheet_name: str = "IO表") -> IoProject:
    path = Path(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"工作表不存在: {sheet_name}，现有: {wb.sheetnames}")
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return _parse_legacy_rows(path.stem, rows)


def export_project_to_workbook(project: IoProject) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "IO"
    from .export import rows_io_table

    for r in rows_io_table(project):
        ws.append(r)
    ws2 = wb.create_sheet("符号表")
    from .export import rows_symbol_table

    for r in rows_symbol_table(project):
        ws2.append(r)
    return wb


def save_project_excel(project: IoProject, path: str | Path) -> None:
    path = Path(path)
    wb = export_project_to_workbook(project)
    wb.save(path)


def import_io_sheet(path: str | Path, sheet_name: str = "IO表") -> IoProject:
    path = Path(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"工作表不存在: {sheet_name}，现有: {wb.sheetnames}")
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return project_single_channel(path.stem, "通道1", [])
    m0 = _map_headers(rows[0])
    if _has_omron_header(m0):
        return _parse_flat_rows(path.stem, rows, m0)
    if _looks_like_legacy(rows):
        return _parse_legacy_rows(path.stem, rows)
    return _parse_fixed_columns(path.stem, rows)


def _parse_legacy_rows(name: str, rows: List) -> IoProject:
    block: Optional[str] = None
    points: List[IoPoint] = []

    for row in rows:
        if not row:
            continue
        r0 = _cell_str(row[0]).upper() if row[0] is not None else ""
        if r0 == "IN":
            block = "IN"
            continue
        if r0 == "OUT":
            block = "OUT"
            continue
        if block is None:
            continue
        col = 0
        widx = 0
        while col + 1 < len(row):
            addr = _norm_addr(row[col])
            comment = _cell_str(row[col + 1])
            if addr:
                points.append(
                    IoPoint(
                        name="",
                        data_type="BOOL",
                        address=addr,
                        comment=comment,
                        rack=str(widx),
                        usage=block,
                    )
                )
            widx += 1
            col += 2

    return project_single_channel(name, "导入", points)
