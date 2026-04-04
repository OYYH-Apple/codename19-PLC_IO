# -*- coding: utf-8 -*-
"""One-off: dump Excel structure to UTF-8 file."""
from pathlib import Path
from openpyxl import load_workbook

path = Path(r"E:\大族激光\创豪\创豪ALF IO表20250704.xlsx")
out = Path(__file__).resolve().parent.parent / "docs" / "reference_excel_dump.txt"
out.parent.mkdir(parents=True, exist_ok=True)

lines = []
wb = load_workbook(path, read_only=True, data_only=True)
lines.append(f"Sheets: {wb.sheetnames}\n")

for name in wb.sheetnames:
    ws = wb[name]
    lines.append(f"\n======== {name} ========\n")
    rows = list(ws.iter_rows(max_row=25, values_only=True))
    for r in rows:
        lines.append(str(r) + "\n")

out.write_text("".join(lines), encoding="utf-8")
print(f"Wrote {out}")
